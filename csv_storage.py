import csv
import os
from datetime import datetime

HEADERS = [
    "Name",
    "Email",
    "Phone",
    "College",
    "Degree",
    "CGPA",
    "Status" 
]


def save_to_csv(details, file_path="resumes.csv"):
    """Save resume details to CSV with duplicate checking"""

    file_exists = os.path.isfile(file_path)

    if file_exists and details.get("email"):
        if is_duplicate_email(details["email"], file_path):
            print(f"⚠️ Duplicate email found: {details['email']}")
            return False

    # Check if headers exist
    headers_missing = True
    if file_exists:
        with open(file_path, "r", encoding="utf-8") as f:
            first_line = f.readline().strip()
            if first_line and "Name" in first_line and "Email" in first_line:
                headers_missing = False

    # Write to CSV
    with open(file_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # Write headers if needed
        if not file_exists or headers_missing:
            writer.writerow(HEADERS)

        # Write data row
        writer.writerow([
            details.get("name", "N/A"),
            details.get("email", "N/A"),
            details.get("phone", "N/A"),
            details.get("college", "N/A"),
            details.get("degree", "N/A"),
            details.get("cgpa", "N/A"),
            "Pending" 
        ])

    print(f"✅ Saved to {file_path}")
    return True


def is_duplicate_email(email, file_path="resumes.csv"):
    """Check if email already exists in CSV"""
    if not os.path.isfile(file_path):
        return False

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get("Email", "").lower() == email.lower():
                    return True
    except Exception as e:
        print(f"⚠️ Error checking duplicates: {e}")

    return False


def get_all_resumes(file_path="resumes.csv"):
    """Read all resumes from CSV"""
    resumes = []

    if not os.path.isfile(file_path):
        return resumes

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                resumes.append(row)
    except Exception as e:
        print(f"❌ Error reading CSV: {e}")

    return resumes


def update_status(email, new_status, file_path="resumes.csv"):
    """Update the status of a resume by email"""
    if not os.path.isfile(file_path):
        return False

    try:
        # Read all rows
        rows = []
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            fieldnames = reader.fieldnames
            for row in reader:
                if row.get("Email", "").lower() == email.lower():
                    row["Status"] = new_status
                rows.append(row)

        # Write back
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

        print(f"✅ Updated status for {email} to {new_status}")
        return True

    except Exception as e:
        print(f"❌ Error updating status: {e}")
        return False


def search_by_cgpa(min_cgpa, file_path="resumes.csv"):
    """Find candidates with CGPA >= min_cgpa"""
    candidates = []

    if not os.path.isfile(file_path):
        return candidates

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                cgpa_str = row.get("CGPA", "0")
                # Extract first number from "9.09 / 10" format
                try:
                    cgpa = float(cgpa_str.split("/")[0].strip())
                    if cgpa >= min_cgpa:
                        candidates.append(row)
                except ValueError:
                    continue
    except Exception as e:
        print(f"❌ Error searching by CGPA: {e}")

    return candidates


def export_to_excel(csv_file_path="resumes.csv", output_path="resumes.xlsx"):
    """Export CSV to Excel format with formatting"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        resumes = get_all_resumes(csv_file_path)

        if not resumes:
            print("⚠️ No resumes to export")
            return False

        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumes"

        # Get headers
        headers = list(resumes[0].keys())

        # Write headers with styling
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Write data
        for row_num, resume in enumerate(resumes, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = resume.get(header, "N/A")
                cell.border = border
                cell.alignment = Alignment(vertical='center')

                # Color code status
                if header == "Status":
                    status = resume.get(header, "Pending")
                    if status == "Reviewed":
                        cell.fill = PatternFill(
                            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100", bold=True)
                    elif status == "Shortlisted":
                        cell.fill = PatternFill(
                            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C6500", bold=True)
                    elif status == "Rejected":
                        cell.fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006", bold=True)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save workbook
        wb.save(output_path)
        print(f"✅ Exported to {output_path}")
        return True

    except ImportError:
        print("❌ openpyxl not installed. Install with: pip install openpyxl")
        return False
    except Exception as e:
        print(f"❌ Error exporting to Excel: {e}")
        return False
